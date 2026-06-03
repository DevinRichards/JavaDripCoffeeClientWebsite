from __future__ import annotations

import re
import sqlite3
import sys
from collections import OrderedDict, defaultdict
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_CATEGORY_SUBTITLES = {
    "Refreshers": "Fruit-forward iced energy and tea refreshers",
    "Lattes": "Espresso and milk in small, medium, and large pours",
    "Reg Coffee": "Daily coffee service with classic sizing",
    "Lemonades": "Bright citrus drinks for an all-day lift",
    "Teas": "Hot tea pours and comforting cafe staples",
    "Matcha": "Matcha-based drinks with layered sweetness",
    "Chai": "Spiced chai drinks across the full size run",
    "Roadrunner": "House specialty drinks with an extra charge of energy",
    "Add Ons": "Optional extras and boosts available at the counter",
    "Milk Options": "Alternative milk upgrades and substitutions",
    "Cake Pops": "Sweet grab-and-go bakery treats",
    "Uncategorized": "Imported from Clover and awaiting a category assignment",
}


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"(^-|-$)", "", value)
    return value or "menu-category"


def unique_slug(name: str, used: set[str]) -> str:
    base = slugify(name)
    candidate = base
    index = 2
    while candidate in used:
        candidate = f"{base}-{index}"
        index += 1
    used.add(candidate)
    return candidate


def load_rows(worksheet):
    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    rows = []
    for row_index in range(2, worksheet.max_row + 1):
        values = [worksheet.cell(row_index, col).value for col in range(1, worksheet.max_column + 1)]
        rows.append(dict(zip(headers, values)))
    return rows


def to_price(value):
    if value in (None, ""):
        return None
    return float(value)


def main():
    if len(sys.argv) != 3:
        print("Usage: import_clover_inventory.py <xlsx-path> <sqlite-db-path>", file=sys.stderr)
        raise SystemExit(1)

    workbook_path = Path(sys.argv[1]).expanduser().resolve()
    database_path = Path(sys.argv[2]).expanduser().resolve()

    wb = load_workbook(workbook_path, data_only=True)
    items_rows = load_rows(wb["Items"])
    category_rows = load_rows(wb["Categories"]) if "Categories" in wb.sheetnames else []

    category_item_order = defaultdict(list)
    category_sheet_order = []
    current_category_name = None

    for row in category_rows:
        category_name = row.get("Category Name") or current_category_name
        item_name = row.get("Item Sort Order")
        if category_name and category_name != current_category_name:
            category_sheet_order.append(category_name)
            current_category_name = category_name
        if category_name and item_name:
            category_item_order[category_name].append(str(item_name).strip())

    items = []
    discovered_categories = OrderedDict()

    for row in items_rows:
        name = (row.get("Name") or "").strip() if row.get("Name") else ""
        if not name:
            continue

        category_name = (row.get("Categories") or "").strip() or "Uncategorized"
        discovered_categories.setdefault(category_name, None)

        price = to_price(row.get("Price"))
        if price is None:
            # Gift cards and blank-price exports are not actionable menu items for the website.
            continue

        items.append(
            {
                "clover_item_id": str(row.get("Clover ID") or "").strip() or None,
                "name": name,
                "description": (row.get("Description") or "").strip() or None,
                "price": price,
                "category_name": category_name,
                "active": 0 if str(row.get("Hidden?") or "").strip().lower() == "yes" else 1,
            }
        )

    ordered_category_names = []
    seen = set()
    for name in category_sheet_order + list(discovered_categories.keys()):
        if name not in seen:
            ordered_category_names.append(name)
            seen.add(name)

    used_slugs = set()
    categories = []
    category_slug_map = {}
    for index, category_name in enumerate(ordered_category_names, start=1):
        slug = unique_slug(category_name, used_slugs)
        category_slug_map[category_name] = slug
        categories.append(
            {
                "id": slug,
                "name": category_name,
                "subtitle": DEFAULT_CATEGORY_SUBTITLES.get(category_name),
                "sort_order": index * 10,
            }
        )

    named_order_lookup = {
        category_name: {item_name: (i + 1) * 10 for i, item_name in enumerate(item_names)}
        for category_name, item_names in category_item_order.items()
    }
    category_item_counters = defaultdict(int)

    for item in items:
        category_name = item["category_name"]
        category_slug = category_slug_map[category_name]
        item["category_id"] = category_slug
        explicit_order = named_order_lookup.get(category_name, {}).get(item["name"])
        if explicit_order is not None:
            item["sort_order"] = explicit_order
        else:
            category_item_counters[category_slug] += 1
            item["sort_order"] = (len(named_order_lookup.get(category_name, {})) + category_item_counters[category_slug]) * 10

    connection = sqlite3.connect(database_path)
    try:
        cursor = connection.cursor()
        cursor.execute("PRAGMA foreign_keys = ON")
        cursor.execute("BEGIN")
        cursor.execute("DELETE FROM menu_items")
        cursor.execute("DELETE FROM menu_categories")

        cursor.executemany(
            """
            INSERT INTO menu_categories (id, name, subtitle, sort_order, clover_category_id)
            VALUES (?, ?, ?, ?, NULL)
            """,
            [(cat["id"], cat["name"], cat["subtitle"], cat["sort_order"]) for cat in categories],
        )

        cursor.executemany(
            """
            INSERT INTO menu_items (
              category_id, name, description, price, image_url, badge, active, sort_order, clover_item_id
            )
            VALUES (?, ?, ?, ?, NULL, NULL, ?, ?, ?)
            """,
            [
                (
                    item["category_id"],
                    item["name"],
                    item["description"],
                    item["price"],
                    item["active"],
                    item["sort_order"],
                    item["clover_item_id"],
                )
                for item in items
            ],
        )

        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()

    print(f"Imported {len(categories)} categories and {len(items)} items from {workbook_path.name}")


if __name__ == "__main__":
    main()
